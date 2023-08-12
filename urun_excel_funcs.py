from openpyxl.styles import Font, PatternFill,  Border, Side, Alignment
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl import load_workbook
import datetime
import openpyxl
from urun_folderjob import clear_files_in_path, copy_file_to_path
from urun_email_content import urun_email_content
from send_email import send_email
from db_funcs import *
from config import *


def format_with_period(n):
    n = str(n)
    n_len = len(n)
    if n_len <= 3:
        return n
    else:
        return format_with_period(n[:-3]) + "." + n[-3:]

def update_excel_file_(marka):
    excel_file = openpyxl.load_workbook('Urun_dpt_rpa_layout.xlsx')
    sheet = excel_file.get_sheet_by_name(marka)

    today = datetime.datetime.today()
    today = today.strftime('%d.%m.%Y')

    if marka == 'DS':
        query = "Select car_detail from [Raporlar].[dbo].[{}_cars_data] where tarih = '{}'".format(marka, today)
    elif marka != 'DS':
        query = "Select car_detail from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
    car_details_list = AssignDBContenttoListWithQuery(YuceDB, query)

    if marka == 'DS':
        query1 = "Select thisyear_prod_price from [Raporlar].[dbo].[{}_cars_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select recommended_sale_price from [Raporlar].[dbo].[{}_cars_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            if fiyat_list1[i] != None:
                fiyat1 = int(fiyat_list1[i])
            else:
                fiyat1 = 0
            if fiyat_list2[i] != None:
                fiyat2 = int(fiyat_list2[i])
            else:
                fiyat2 = 0
            if fiyat1 == 0:
                fiyat = fiyat2
            elif fiyat2 == 0:
                fiyat = fiyat1
            fiyat_list.append(str(fiyat))

    elif marka == 'Ford':
        query1 = "Select T_Edilen_Anahtar_Teslim_fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select T_Edilen_Kampanyali_Anahtar_Teslim_fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Kia':
        query1 = "Select Kampanyali_Satis_Fiyati from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Satis_Fiyati from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Toyota':
        query1 = "Select fiyat1 from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select fiyat2 from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1,fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Renault':
        query1 = "Select buyil_model_fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select kampanyali_fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Honda':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Peugeot':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(str(fiyat_list1[i]).replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(str(fiyat_list2[i]).replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Cupra':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Opel':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0").replace("?",""))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Seat':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(
            marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(
            marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Citroen':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Volkswagen':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Nissan':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Nakit_Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))

    elif marka == 'Hyundai':
        query1 = "Select Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list1 = AssignDBContenttoListWithQuery(YuceDB, query1)
        query2 = "Select Kampanyali_Fiyat from [Raporlar].[dbo].[{}_auto_data] where tarih = '{}'".format(marka, today)
        fiyat_list2 = AssignDBContenttoListWithQuery(YuceDB, query2)
        fiyat_list = []
        for i in range(len(fiyat_list1)):
            fiyat1 = int(fiyat_list1[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            fiyat2 = int(fiyat_list2[i].replace(" TL", "").replace(".", "").replace("-", "0"))
            if fiyat2 != 0:
                fiyat = min(fiyat1, fiyat2)
            else:
                fiyat = fiyat1
            if fiyat == 0:
                fiyat = "-"
            fiyat_list.append(str(fiyat))


    excel_arac_veriler = []
    for i in range(200):
        box = "A" + str(i + 2)
        value = sheet[box].value
        excel_arac_veriler.append(value)

    if ".01." in today:
        tarih = today[:2] + " Ocak " + today[-4:]
    elif ".02." in today:
        tarih = today[:2] + " Şubat " + today[-4:]
    elif ".03." in today:
        tarih = today[:2] + " Mart " + today[-4:]
    elif ".04." in today:
        tarih = today[:2] + " Nisan " + today[-4:]
    elif ".05." in today:
        tarih = today[:2] + " Mayıs " + today[-4:]
    elif ".06." in today:
        tarih = today[:2] + " Haziran " + today[-4:]
    elif ".07." in today:
        tarih = today[:2] + " Temmuz " + today[-4:]
    elif ".08." in today:
        tarih = today[:2] + " Ağustos " + today[-4:]
    elif ".09." in today:
        tarih = today[:2] + " Eylül " + today[-4:]
    elif ".10." in today:
        tarih = today[:2] + " Ekim " + today[-4:]
    elif ".11." in today:
        tarih = today[:2] + " Kasım " + today[-4:]
    elif ".12." in today:
        tarih = today[:2] + " Aralık " + today[-4:]

    new_col = sheet['B1'].value
    new_col_int = 2
    if new_col == None:
        prev_col = "A"
        prev_prev_col = ""
    if new_col != None:
        new_col = sheet['C1'].value
        col_name = "C"
        prev_col = "B"
        prev_prev_col = "A"
        new_col_int = 3
    if new_col != None:
        new_col = sheet['D1'].value
        col_name = "D"
        prev_col = "C"
        prev_prev_col = "B"
        new_col_int = 4
    if new_col != None:
        new_col = sheet['E1'].value
        col_name = "E"
        prev_col = "D"
        prev_prev_col = "C"
        new_col_int = 5
    if new_col != None:
        new_col = sheet['F1'].value
        col_name = "F"
        prev_col = "E"
        prev_prev_col = "D"
        new_col_int = 6
    if new_col != None:
        new_col = sheet['G1'].value
        col_name = "G"
        prev_col = "F"
        prev_prev_col = "E"
        new_col_int = 7
    if new_col != None:
        new_col = sheet['H1'].value
        col_name = "H"
        prev_col = "G"
        prev_prev_col = "F"
        new_col_int = 8
    if new_col != None:
        new_col = sheet['I1'].value
        col_name = "I"
        prev_col = "H"
        prev_prev_col = "G"
        new_col_int = 9
    if new_col != None:
        new_col = sheet['J1'].value
        col_name = "J"
        prev_col = "I"
        prev_prev_col = "H"
        new_col_int = 10
    if new_col != None:
        new_col = sheet['K1'].value
        col_name = "K"
        prev_col = "J"
        prev_prev_col = "I"
        new_col_int = 11
    if new_col != None:
        new_col = sheet['L1'].value
        col_name = "L"
        prev_col = "K"
        prev_prev_col = "J"
        new_col_int = 12
    if new_col != None:
        new_col = sheet['M1'].value
        col_name = "M"
        prev_col = "L"
        prev_prev_col = "K"
        new_col_int = 13
    if new_col != None:
        new_col = sheet['N1'].value
        col_name = "N"
        prev_col = "M"
        prev_prev_col = "L"
        new_col_int = 14
    if new_col != None:
        new_col = sheet['O1'].value
        col_name = "O"
        prev_col = "N"
        prev_prev_col = "M"
        new_col_int = 15
    if new_col != None:
        new_col = sheet['P1'].value
        col_name = "P"
        prev_col = "O"
        prev_prev_col = "N"
        new_col_int = 16
    if new_col != None:
        new_col = sheet['Q1'].value
        col_name = "Q"
        prev_col = "P"
        prev_prev_col = "O"
        new_col_int = 17
    if new_col != None:
        new_col = sheet['R1'].value
        col_name = "R"
        prev_col = "Q"
        prev_prev_col = "P"
        new_col_int = 18
    if new_col != None:
        new_col = sheet['S1'].value
        col_name = "S"
        prev_col = "R"
        prev_prev_col = "Q"
        new_col_int = 19
    if new_col != None:
        new_col = sheet['T1'].value
        col_name = "T"
        prev_col = "S"
        prev_prev_col = "R"
        new_col_int = 20
    if new_col != None:
        new_col = sheet['U1'].value
        col_name = "U"
        prev_col = "T"
        prev_prev_col = "S"
        new_col_int = 21
    if new_col != None:
        new_col = sheet['V1'].value
        col_name = "V"
        prev_col = "U"
        prev_prev_col = "T"
        new_col_int = 22
    if new_col != None:
        new_col = sheet['W1'].value
        col_name = "W"
        prev_col = "V"
        prev_prev_col = "U"
        new_col_int = 23
    if new_col != None:
        new_col = sheet['X1'].value
        col_name = "X"
        prev_col = "W"
        prev_prev_col = "V"
        new_col_int = 24
    if new_col != None:
        new_col = sheet['Y1'].value
        col_name = "Y"
        prev_col = "X"
        prev_prev_col = "W"
        new_col_int = 25

    sheet.insert_cols(new_col_int) # Inserts an empty row after the last column of prices

    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    bold_font = Font(bold=True)

    sheet['{}1'.format(col_name)] = tarih
    sheet['{}1'.format(col_name)].alignment = Alignment(horizontal='center', vertical='center')
    sheet['{}1'.format(col_name)].font = bold_font
    sheet['{}1'.format(col_name)].border = border_style

    # Dyes the background to dark red of the Date added, in 1st row
    dark_red = "7A1300"
    fill = PatternFill(start_color=dark_red, end_color=dark_red, fill_type="solid")
    sheet['{}1'.format(col_name)].fill = fill

    # Makes the text-color white
    white_text_color = "FFFFFFFF"
    font2 = Font(color=white_text_color, bold=True)
    sheet['{}1'.format(col_name)].font = font2


    sheet.column_dimensions[col_name].width = 16
    amount = 0
    for i in range(len(car_details_list)):
        if car_details_list[i] in excel_arac_veriler: # IF THE MODEL EXISTS IN THE EXCEL SHEET
            line = excel_arac_veriler.index(car_details_list[i])
            box = col_name + str(line + 2)
            try:
                fiyat_list[i] = fiyat_list[i].replace("TL", "").replace(" ", "").replace(".", "")
            except:pass
            fiyat_list[i] = format_with_period(fiyat_list[i])
            sheet[box] = fiyat_list[i]
            cell = sheet[box]
            cell.border = border_style
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else: # IF THE MODEL DOES NOT EXIST IN THE EXCEL SHEET
            sequence_of_none = 0
            count = 3
            for veri in excel_arac_veriler:
                count += 1
                if veri == None:
                    sequence_of_none += 1
                    if sequence_of_none == 2:
                        line_to_be_written = count + amount
                        amount += 1
                elif veri != None:
                    sequence_of_none = 0
            boxy = 'A' + str(line_to_be_written)

            #sheet['A{}'.format(str(line_to_be_written-1))] = 'Yeni eklenen model:'
            sheet['A{}'.format(str(line_to_be_written - 1))].font = bold_font
            sheet['A{}'.format(str(line_to_be_written - 1))].alignment = Alignment(horizontal='center', vertical='center')

            sheet[boxy] = car_details_list[i]
            sheet[boxy].font = bold_font
            sheet[boxy].alignment = Alignment(horizontal='center', vertical='center')

            fiyat_to_be_written = col_name + str(line_to_be_written)
            sheet[fiyat_to_be_written] = format_with_period(fiyat_list[i])
            sheet[fiyat_to_be_written].font = bold_font
            sheet[fiyat_to_be_written].alignment = Alignment(horizontal='center', vertical='center')
            sheet[fiyat_to_be_written].border = border_style

            for letter in range(ord('A'), ord('Z') + 1):
                column = chr(letter)
                if column != 'A':
                    position = column + str(line_to_be_written)
                    sheet[position] = '-'
                    sheet[position].font = bold_font
                    sheet[position].alignment = Alignment(horizontal='center', vertical='center')
                    sheet[position].border = border_style
                if prev_col == chr(letter):
                    break

    prev_column_values = []
    for i in range(100):
        prev_vals = prev_col + str(i + 2)
        prev_column_values.append(sheet[prev_vals].value)

    today_column_values = []
    for i in range(100):
        today_vals = col_name + str(i + 2)
        today_column_values.append(sheet[today_vals].value)

    Annual_col = None
    for column in sheet.iter_cols():
        if column[0].value == 'Yıllık Değişim':
            Annual_col = column[0].column_letter
            break

    # Iterates through each row and writes the annual change percentage to the "Yıllık Değişim" column on each row possible
    for row in range(2, 201):
        column_b_value = sheet.cell(row=row, column=2).value
        column_new_value = sheet.cell(row=row, column=column_index_from_string(col_name)).value
        if column_b_value != None:
            if type(column_b_value) != int:
                if column_b_value != "-":
                    column_b_value = int(column_b_value.replace(".","").replace(" ",""))
                else:
                    column_b_value = 0
        else:
            column_b_value = 0
        if column_new_value != None:
            if type(column_new_value) != int:
                if column_new_value != "-":
                    column_new_value = int(column_new_value.replace(".", "").replace(" ", ""))
                else:
                    column_new_value = 0
        else:
            column_new_value = 0

        if column_b_value == 0 or column_new_value == 0:
            price_change = "-"
        else:
            price_change = str(round((((column_new_value - column_b_value) / column_b_value) * 100),2)) + "%"

            sheet[Annual_col + str(row)] = str(price_change)
            sheet[Annual_col + str(row)].font = bold_font
            sheet[Annual_col + str(row)].alignment = Alignment(horizontal='center', vertical='center')
            sheet[Annual_col + str(row)].border = border_style

    sum_values = 0
    count = 0
    for cell in sheet[Annual_col]:
        if cell.value != None:
            change = cell.value.replace("%","")
            try:
                sum_values = sum_values + float(change)
                count += 1
            except:pass
    avg_price_inc = round((sum_values / count), 2)
    sheet[Annual_col + "2"] = str(avg_price_inc).replace(".",",") + "%"


    # If any price difference is spotted, then it will save the excel, else it will not
    for i in range(len(today_column_values)):
        if today_column_values[i] == None:
            today_column_values[i] = '-'
        if prev_column_values[i] == None:
            prev_column_values[i] = '-'

    empty_column_count = 0
    column_count = 0

    for column in sheet.iter_cols(values_only=True):
        column_count += 1
        if all(cell is None for cell in column):
            empty_column_count += 1
            if empty_column_count == 2:
                second_empty_column_letter = get_column_letter(column_count)
                break

    # Calculates the price changes of the last 2 column, last column existing and today's prices
    percentages = []
    for g in range(len(today_column_values)):
        if today_column_values[g] != "-":
            try:
                today_price = int(today_column_values[g])
            except:
                today_price = int(today_column_values[g].replace(".", "").replace(" ", ""))
        else:
            today_price = "-"
        if prev_column_values[g] != "-":
            try:
                last_price = int(prev_column_values[g])
            except:
                last_price = int(prev_column_values[g].replace(".", "").replace(" ", ""))
        else:
            last_price = "-"
        try:
            if today_price != "-" and last_price != "-":
                percentage = ((today_price - last_price) / last_price) * 100
            else:
                percentage = "+"
        except:
            percentage = "+"
        try:
            percentage = round(percentage, 2)
        except:
            percentage = "+"
        percentages.append(str(percentage) + "%")

    l = 0
    for item in reversed(excel_arac_veriler):
        l += 1
        if item != None:
            f = 200-l
            break
    nones_in_arac_veriler = []
    for i in range(f):
        if excel_arac_veriler[i] == None:
            nones_in_arac_veriler.append(i + 1)

    l = 0
    avg_value_list = []
    for gh in range(len(percentages)):
        l += 1
        if percentages[gh] != "+%":
            try:
                avg_value_list.append(float(percentages[gh].replace("%", '')))
            except:pass
            sheet[second_empty_column_letter + str(gh + 2)] = percentages[gh]
            sheet[second_empty_column_letter + str(gh + 2)].font = bold_font
            sheet[second_empty_column_letter + str(gh + 2)].alignment = Alignment(horizontal='center', vertical='center')
            sheet[second_empty_column_letter + str(gh + 2)].border = border_style
        else:
            if l <= f + 1:
                if l not in nones_in_arac_veriler:
                    sheet[second_empty_column_letter + str(gh + 2)] = "-"
                    sheet[second_empty_column_letter + str(gh + 2)].font = bold_font
                    sheet[second_empty_column_letter + str(gh + 2)].alignment = Alignment(horizontal='center', vertical='center')
                    sheet[second_empty_column_letter + str(gh + 2)].border = border_style
    if len(avg_value_list) != 0:
        average = round(sum(avg_value_list) / len(avg_value_list), 2)
    else:
        average = 0

    sheet[second_empty_column_letter + "2"] = str(average) + "%"
    sheet[second_empty_column_letter + "2"].font = bold_font
    sheet[second_empty_column_letter + "2"].alignment = Alignment(horizontal='center', vertical='center')
    sheet[second_empty_column_letter + "2"].border = border_style

    # If today_column_values have list of '-' in first 15 elements, then it's empty
    column_empty = False
    if today_column_values[0] == '-' and today_column_values[1] == '-' and today_column_values[2] == '-' and today_column_values[3] == '-' and today_column_values[4] == '-' and today_column_values[5] == '-' and today_column_values[6] == '-' and today_column_values[7] == '-' and today_column_values[8] == '-' and today_column_values[9] == '-' and today_column_values[10] == '-' and today_column_values[11] == '-' and today_column_values[12] == '-' and today_column_values[13] == '-':
        column_empty = True

    # If last day prices and today entered prices have any difference, if condition returns true
    if prev_column_values != today_column_values and column_empty == False:
        sondan_ikinci_column = sheet['{}1'.format(prev_col)].value
        sondan_birinci_column = sheet['{}1'.format(col_name)].value
        red_box_to_be_filled = "{}1".format(second_empty_column_letter)
        sheet[red_box_to_be_filled] = sondan_ikinci_column[:6] + " - " + sondan_birinci_column[:6]


        # Column width of the first 50 columns is declared here (First one skipped)
        column_width = 16
        for column_index in range(2, 51):
            column_letter = get_column_letter(column_index)
            sheet.column_dimensions[column_letter].width = column_width

        # Dyes 2nd row to electric green
        electric_green = "78FAAE"
        fill = PatternFill(start_color=electric_green, end_color=electric_green, fill_type="solid")
        for cell in sheet[2]:
            cell.fill = fill

        # Dyes the even numbered rows to light grey
        Light_grey = "E3E5E6"
        fill = PatternFill(start_color=Light_grey, end_color=Light_grey, fill_type="solid")
        for row in range(4, 201, 2):
            for cell in sheet[row]:
                cell.fill = fill

        # Dyes the empty rows to dark Emerald Green
        emerald_green = "0E3A2F"
        fill = PatternFill(start_color=emerald_green, end_color=emerald_green, fill_type="solid")
        for row_num in range(3, 101):
            cell_value = sheet.cell(row=row_num, column=1).value
            for cell in sheet[row]:
                cell.fill = fill
            if cell_value is None:
                for cell in sheet[row_num]:
                    cell.fill = fill

        new_empty_column = new_col_int + 1
        for row_num in range(1, 101):
            cell = sheet.cell(row=row_num, column=new_empty_column)
            cell.fill = fill


        excel_file.save('Urun_dpt_rpa_layout.xlsx')

        # IF CHANGES ARE DETECTED,
        clear_files_in_path(r'Y:\YUCE AUTO GENEL\RPA\Markalar_Fiyat_Güncelleme')
        copy_file_to_path('Urun_dpt_rpa_layout.xlsx', r'Y:\YUCE AUTO GENEL\RPA\Markalar_Fiyat_Güncelleme')

        email_content = urun_email_content.format(marka)
        receiver = "b.yurdasiper@skoda.com.tr; afizeo@skoda.com.tr;"
        send_email(email_content, 'Fiyat Değişikliği', receiver, cc_email=None)
        receivers_urun = "i.tekin@skoda.com.tr; stj_d.keler@skoda.com.tr; y.uzunoglu@skoda.com.tr; e.katiranci@skoda.com.tr;"
        send_email(email_content, 'Fiyat Değişikliği', receivers_urun, cc_email=None)
        send_email(email_content, 'Fiyat Değişikliği', 'l.akkus@skoda.com.tr', cc_email= 'dogao@skoda.com.tr')
        print("There have been changes in {}, saved to excel, notified via email....".format(marka))
        return 1
    else:
        return 0
